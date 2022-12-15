from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import time
#font for title
titlefont = Font(bold=True)


wb = Workbook()
###global
i = 1 # a counter which gets reset
item_number = 0 #count the items
item_number_offset = 2 #number of rows before count starts
date = 151
site_project = 'darigold_P21034120'
name = "Grayson Stillwell"
filename = 'EEL'
machine_type = 'SRT'
design_air_temperature_C = 3
total_evaporator_count = 6
total_fan_count = 20
largest_fan_kw = 2.2
VSD_DOL_selection = "VSD"
infeed_conveyor_count = 1
outfeed_conveyor_count = 1

dest_filename = site_project + str(date) + filename +".xlsx"
ws1 = wb.active
ws1.title = "TitlePage"
ws1['A1'] = "MHM"
ws1['A1'].font = titlefont  # Bold
ws1['B1'] = "Machine Type"
ws1['B2'] = machine_type
ws1['C1'] = "site + ProjectNumber"
ws1['C2'] = site_project
ws1['D1'] = "Total Number Of Fans"
ws1['D2'] = total_fan_count
ws1['E1'] = "Evaporator Count"
ws1['E2'] = total_evaporator_count
ws1['F1'] = "Largest KW rating of fan motor"
ws1['F2'] = largest_fan_kw
ws1['G1'] = "Fans, VSD or DOL"
ws1['G2'] = VSD_DOL_selection
ws1['H1'] = "Designed Air Temperature C, [NOTE: this determines dehumidifier & carton stop] "
ws1['H2'] = design_air_temperature_C
ws1['I1'] = "Infeed Conveyor Count"
ws1['I2'] = infeed_conveyor_count
ws1['J1'] = "Outfeed Conveyor Count"
ws1['J2'] = outfeed_conveyor_count

ws2 = wb.create_sheet(title="Device Schedule")
ws2['A1'] = "MHM"
ws2['A2'] = "Item #"
ws2['B2'] = "Change Log"	
ws2['C2'] = "Change Date"
ws2['D2'] = "Area"	
ws2['E2'] = "Equipment ID"	
ws2['F2'] = "Equipment #"	
ws2['G2'] = "Device ID"	
ws2['H2'] = "Device #"	
ws2['I2'] = "Device Tag"		
ws2['J2'] = "Device Description"
ws2['K2'] = "Device Description 2"	
ws2['L2'] = "Design Notes"	
ws2['M2'] = "Location / Origin (F)" 
ws2['N2'] = "Field"	
ws2['O2'] = "Rated kW"	
ws2['P2'] = "Load Type"	
ws2['Q2'] = "E-Stop Zone"	
ws2['R2'] = "Allocated Address (PLC Format)"	
ws2['S2'] = "Input / Output"	
ws2['T2'] = "Ethernet"	
ws2['U2'] = "IO-Link DI"	
ws2['V2'] = "IO-Link DO"	
ws2['W2'] = "Point IO DI"	
ws2['X2'] = "Point IO DO"
ws2['Y2'] = "Point IO Safe DI"
ws2['Z2'] = "Point IO Safe DO"
ws2['AA2'] = "Point IO AI"	
ws2['AB2'] = "Point IO AO"	
ws2['AC2'] = "Local DI"	
ws2['AD2'] = "Local DO"	
ws2['AE2'] = "Local AI"	
ws2['AF2'] = "Local AO"	
ws2['AG2'] = "Local HSC"	
ws2['AH2'] = "Procurement Status"	
ws2['AI2'] = "Part #2 (Sensor)"
ws2['AJ2'] = "Part #2 (Mounting)"	
ws2['AK2'] = "Part #3" 	
ws2['AL2'] = "Part #4 (Cable/Plug)"	
ws2['AM2'] = "Part #5 (Cable/Plug)"	
ws2['AN2'] = "Comments"	
ws2['AO2'] = "Part #1 Protection"	
ws2['AP2'] = "Part #2 Switchgear"	
ws2['AQ2'] = "Part #3" 
ws2['AR2'] = "Part #4" 	
ws2['AS2'] = "Part #5     (Enclosure)"
ws2['AT2'] = "Comments"
ws2['AU2'] = "Part #1      (Cable Marker)"																																					
ws2['AV2'] = "Part #2       (Cable Type)"
ws2['AW2'] = "Part #3     (Cable Length)" 
ws2['AX2'] = "Conductor Size"
ws2['AY2'] = "Part #4      (Motor Isolator)"
ws2['AZ2'] = "Part #5     (Enclosure)"
ws2['BA2'] = "Comments"


###EVAPORATOR LOCAL SETTINGS
localname = "EVAPORATOR"
Area = "CTR"
EquipmentID = "EV0"
DeviceID = "TT0"
DeviceDescription = "Tunnel Evaporator Temperatures"
###EVAPORATOR TITTLE
ws2.append ([item_number,localname])
item_number += 1
titlelocation = item_number + item_number_offset
fontchange = 'B'+str(titlelocation)
cell = ws2[fontchange]
cell.font = Font(name='Arial',
                 size=12,
                 bold=True,
                 italic=True,
                 underline='single'
                 )
cell.fill = PatternFill("solid", start_color="FFA500")

###EVAPORATOR DATA CREATION
n = 51
while i < total_evaporator_count:
      localname = {
"Item" : item_number,	
"ChangeLog" : name,	
"ChangeDate" : datetime.now(),
"Area" : Area,	
"EquipmentID" : EquipmentID,
"Equipment" : i,
"DeviceID" : DeviceID+str(i),
"Device" : EquipmentID+str(i),	
"DeviceTag" : Area+'_'+EquipmentID+str(i)+'_'+DeviceID+str(i),	
"DeviceDescription" : DeviceDescription,	
"DeviceDescription2" : "",	
"DesignNotes" : "IO-Link, 20m Max Cable Length",
"LocationOriginField" : "CP02",
"RatedkW" : "",	
"LoadType" : "",	
"EStopZone" : "",	
"AllocatedAddressPLC_Format" : "",	
"InputOutput" : "IO-Link",	
"Ethernet" : "",	
"IOLinkDI" : "1", 
"IOLinkDO" : "",	
"PointIODI" : "",	
"PointIODO" : "",
"PointIOSafeDI" : "",	
"PointIOSafeDO" : "",	
"PointIOAI" : "",	
"PointIOAO" : "",	
"LocalDI" : "",	
"LocalDO" : "",	
"LocalAI" : "",
"LocalAO" : "",	
"LocalHSCProcurementStatus" : "",	
"Part1_Sensor" : "",	
"Part2_Mounting" : "",	
"Part3_": "",
"Part_4_CablePlug": "",
"Part_5_CablePlug": "",
"Comments1" : "",
"Part1_Protection":"",
"Part2_Switchgear": "",
"Part3_" : "",	
"Part4_" : "",	
"Part5_Enclosure" : "",
"Comments2" : "",	
"Part1_CableMarker" : "",
"Part2_CableType" : "",
"Part3_CableLength" : "",
"Conductor_Size" : "",
"Part4_MotorIsolator" : "",	
"Part5_Enclosure" : "",	
"Comments3" : "",

}
      values = list(localname.values())[:n]
      i += 1
      item_number += 1
      ws2.append(values)






###FANS LOCAL SETTINGS
i = 1 # a counter which gets reset
localname = "FAN"
Area = "CTR"
EquipmentID = "FN0"
DeviceID = "TT0"
DeviceDescription = ""
###FANS TITTLE
ws2.append ([item_number,localname])
item_number += 1
titlelocation = item_number + item_number_offset
fontchange = 'B'+str(titlelocation)
cell = ws2[fontchange]
cell.font = Font(name='Arial',
                 size=12,
                 bold=True,
                 italic=True,
                 underline='single'
                 )
cell.fill = PatternFill("solid", start_color="FFA500")

###FANS DATA CREATION
n = 51
while i < total_fan_count:
      localname = {
"Item" : item_number,	
"ChangeLog" : name,	
"ChangeDate" : datetime.now(),
"Area" : Area,	
"EquipmentID" : EquipmentID,
"Equipment" : i,
"DeviceID" : DeviceID+str(i),
"Device" : EquipmentID+str(i),	
"DeviceTag" : Area+'_'+EquipmentID+str(i)+'_'+DeviceID+str(i),	
"DeviceDescription" : "Tunnel Evaporator Temperatures",	
"DeviceDescription2" : "",	
"DesignNotes" : "IO-Link, 20m Max Cable Length",
"LocationOriginField" : "CP02",
"RatedkW" : "",	
"LoadType" : "",	
"EStopZone" : "",	
"AllocatedAddressPLC_Format" : "",	
"InputOutput" : "IO-Link",	
"Ethernet" : "",	
"IOLinkDI" : "1", 
"IOLinkDO" : "",	
"PointIODI" : "",	
"PointIODO" : "",
"PointIOSafeDI" : "",	
"PointIOSafeDO" : "",	
"PointIOAI" : "",	
"PointIOAO" : "",	
"LocalDI" : "",	
"LocalDO" : "",	
"LocalAI" : "",
"LocalAO" : "",	
"LocalHSCProcurementStatus" : "",	
"Part1_Sensor" : "",	
"Part2_Mounting" : "",	
"Part3_": "",
"Part_4_CablePlug": "",
"Part_5_CablePlug": "",
"Comments1" : "",
"Part1_Protection":"",
"Part2_Switchgear": "",
"Part3_" : "",	
"Part4_" : "",	
"Part5_Enclosure" : "",
"Comments2" : "",	
"Part1_CableMarker" : "",
"Part2_CableType" : "",
"Part3_CableLength" : "",
"Conductor_Size" : "",
"Part4_MotorIsolator" : "",	
"Part5_Enclosure" : "",	
"Comments3" : "",

}
      values = list(localname.values())[:n]
      i += 1
      item_number += 1
      ws2.append(values)




###HYDRAULICS
i = 1 # a counter which gets reset
localname = "HYDRAULICS"
Area = "CTH"
EquipmentID = "HPU"
DeviceID = "AL"
AllocationID = i #if i = 1 first IO = 1
iomaster_count_total = 3
iomaster_count_start = 1
###HYDRAULICS
ws2.append ([item_number,localname])
item_number += 1
titlelocation = item_number + item_number_offset
fontchange = 'B'+str(titlelocation)
cell = ws2[fontchange]
cell.font = Font(name='Arial',
                 size=12,
                 bold=True,
                 italic=True,
                 underline='single'
                 )
cell.fill = PatternFill("solid", start_color="FFA500")

###HYDRAULICS DATA CREATION
## local IO master
while i <= iomaster_count_total:
      localname = {
"Item" : item_number,	
"ChangeLog" : name,	
"ChangeDate" : datetime.now(),
"Area" : Area,	
"EquipmentID" : EquipmentID,
"Equipment" : i,
"DeviceID" : DeviceID,
"Device" : EquipmentID+str(i),	
"DeviceTag" : Area+'_'+EquipmentID+'_'+DeviceID+"_X0"+str(i),	
"DeviceDescription" : "Tunnel Evaporator Temperatures",	
"DeviceDescription2" : "",	
"DesignNotes" : "IO-Link, 20m Max Cable Length",
"LocationOriginField" : "CP02",
"RatedkW" : "",	
"LoadType" : "",	
"EStopZone" : "",	
"AllocatedAddressPLC_Format" : "",	
"InputOutput" : "IO-Link",	
"Ethernet" : "",	
"IOLinkDI" : "1", 
"IOLinkDO" : "",	
"PointIODI" : "",	
"PointIODO" : "",
"PointIOSafeDI" : "",	
"PointIOSafeDO" : "",	
"PointIOAI" : "",	
"PointIOAO" : "",	
"LocalDI" : "",	
"LocalDO" : "",	
"LocalAI" : "",
"LocalAO" : "",	
"LocalHSCProcurementStatus" : "",	
"Part1_Sensor" : "",	
"Part2_Mounting" : "",	
"Part3_": "",
"Part_4_CablePlug": "",
"Part_5_CablePlug": "",
"Comments1" : "",
"Part1_Protection":"",
"Part2_Switchgear": "",
"Part3_" : "",	
"Part4_" : "",	
"Part5_Enclosure" : "",
"Comments2" : "",	
"Part1_CableMarker" : "",
"Part2_CableType" : "",
"Part3_CableLength" : "",
"Conductor_Size" : "",
"Part4_MotorIsolator" : "",	
"Part5_Enclosure" : "",	
"Comments3" : "",

}
      values = list(localname.values())[:n]
      i += 1
      item_number += 1
      ws2.append(values)
## IO master allocation











     

wb.save(filename = dest_filename)      # save work